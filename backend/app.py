import io
import os
import threading
import time
from datetime import datetime
from pathlib import Path

# Load .env file if present
env_path = Path(__file__).resolve().parent.parent / ".env"
if env_path.exists():
    for line in env_path.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            key, val = line.split("=", 1)
            os.environ.setdefault(key.strip(), val.strip())

from flask import Flask, jsonify, request, send_from_directory, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database import init_db, get_db, get_setting, set_setting, get_keywords, add_keyword, delete_keyword, save_position, get_latest_positions, get_history
from scraper import check_keyword, delay_between_keywords

app = Flask(__name__, static_folder=os.path.join(os.path.dirname(__file__), '..', 'frontend'))

# Initialize database on import
init_db()

# Scan state
scan_state = {
    "running": False,
    "current_keyword": None,
    "current_index": 0,
    "total": 0,
    "started_at": None,
    "next_check_at": None,
}
scan_lock = threading.Lock()


@app.route("/")
def index():
    return send_from_directory(app.static_folder, "index.html")


# --- Settings ---

@app.route("/api/settings", methods=["GET"])
def api_get_settings():
    domain = get_setting("target_domain") or ""
    return jsonify({"target_domain": domain})


@app.route("/api/settings", methods=["POST"])
def api_set_settings():
    data = request.get_json()
    domain = data.get("target_domain", "").strip()
    if not domain:
        return jsonify({"error": "Domain is required"}), 400
    set_setting("target_domain", domain)
    return jsonify({"ok": True})


# --- Keywords ---

@app.route("/api/keywords", methods=["GET"])
def api_get_keywords():
    return jsonify(get_keywords())


@app.route("/api/keywords", methods=["POST"])
def api_add_keyword():
    data = request.get_json()
    # Support bulk: accept "keywords" (list) or "keyword" (single string)
    keywords_list = data.get("keywords", [])
    single = data.get("keyword", "").strip()
    if single and not keywords_list:
        keywords_list = [single]

    if not keywords_list:
        return jsonify({"error": "Keyword is required"}), 400

    added = []
    skipped = []
    for kw in keywords_list:
        kw = kw.strip()
        if not kw:
            continue
        kid = add_keyword(kw)
        if kid is None:
            skipped.append(kw)
        else:
            added.append({"id": kid, "keyword": kw})

    return jsonify({"added": added, "skipped": skipped, "count": len(added)}), 201


@app.route("/api/keywords/clear", methods=["POST"])
def api_clear_keywords():
    """Delete all keywords."""
    keywords = get_keywords()
    for kw in keywords:
        delete_keyword(kw["id"])
    return jsonify({"ok": True, "deleted": len(keywords)})


@app.route("/api/keywords/<int:kid>", methods=["DELETE"])
def api_delete_keyword(kid):
    delete_keyword(kid)
    return jsonify({"ok": True})


# --- Scan ---

def run_scan():
    with scan_lock:
        if scan_state["running"]:
            return
        scan_state["running"] = True

    try:
        domain = get_setting("target_domain")
        if not domain:
            return

        keywords = get_keywords()
        scan_state["total"] = len(keywords)
        scan_state["started_at"] = datetime.now().isoformat()

        for i, kw in enumerate(keywords):
            scan_state["current_index"] = i
            scan_state["current_keyword"] = kw["keyword"]
            scan_state["next_check_at"] = None

            position, url_found = check_keyword(kw["keyword"], domain)
            save_position(kw["id"], position, url_found)

            # Short delay between API calls
            if i < len(keywords) - 1:
                next_time = time.time() + 3
                scan_state["next_check_at"] = datetime.fromtimestamp(next_time).isoformat()
                delay_between_keywords()
    finally:
        with scan_lock:
            scan_state["running"] = False
            scan_state["current_keyword"] = None
            scan_state["next_check_at"] = None


@app.route("/api/check", methods=["POST"])
def api_check():
    if scan_state["running"]:
        return jsonify({"error": "Scan already running"}), 409
    domain = get_setting("target_domain")
    if not domain:
        return jsonify({"error": "Set a target domain first"}), 400
    keywords = get_keywords()
    if not keywords:
        return jsonify({"error": "Add at least one keyword"}), 400
    t = threading.Thread(target=run_scan, daemon=True)
    t.start()
    return jsonify({"ok": True, "total": len(keywords)})


@app.route("/api/status", methods=["GET"])
def api_status():
    return jsonify({
        "running": scan_state["running"],
        "current_keyword": scan_state["current_keyword"],
        "current_index": scan_state["current_index"],
        "total": scan_state["total"],
        "started_at": scan_state["started_at"],
        "next_check_at": scan_state["next_check_at"],
    })


# --- Results ---

@app.route("/api/results", methods=["GET"])
def api_results():
    return jsonify(get_latest_positions())


@app.route("/api/history/<int:keyword_id>", methods=["GET"])
def api_history(keyword_id):
    return jsonify(get_history(keyword_id))


# --- Excel Export ---

@app.route("/api/export", methods=["GET"])
def api_export():
    domain = get_setting("target_domain") or "unknown"
    results = get_latest_positions()

    wb = Workbook()
    ws = wb.active
    ws.title = "Positions SEO"

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="6C5CE7", end_color="6C5CE7", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color="E8E8EE")
    )
    top3_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    top10_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    top100_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    na_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

    # Title row
    ws.merge_cells("A1:D1")
    ws["A1"] = f"SEO Position Report — {domain}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="6C5CE7")
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    # Date row
    ws.merge_cells("A2:D2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri", size=10, color="6B6E7B")
    ws.row_dimensions[2].height = 20

    # Empty row
    ws.row_dimensions[3].height = 10

    # Headers
    headers = ["Mot-clé", "Position", "URL trouvée", "Date vérification"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.row_dimensions[4].height = 28

    # Data rows
    for i, r in enumerate(results, 5):
        ws.cell(row=i, column=1, value=r["keyword"]).border = thin_border
        pos_cell = ws.cell(row=i, column=2)
        pos_cell.border = thin_border
        pos_cell.alignment = Alignment(horizontal="center")

        if r["position"] is not None:
            pos_cell.value = r["position"]
            if r["position"] <= 3:
                pos_cell.fill = top3_fill
                pos_cell.font = Font(bold=True, color="155724")
            elif r["position"] <= 10:
                pos_cell.fill = top10_fill
                pos_cell.font = Font(bold=True, color="856404")
            else:
                pos_cell.fill = top100_fill
                pos_cell.font = Font(bold=True, color="721C24")
        else:
            pos_cell.value = "Non trouvé"
            pos_cell.fill = na_fill
            pos_cell.font = Font(color="6B6E7B")

        ws.cell(row=i, column=3, value=r["url_found"] or "—").border = thin_border
        ws.cell(row=i, column=4, value=r["checked_at"] or "Jamais").border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 22

    # Summary row
    row_sum = len(results) + 6
    found = [r for r in results if r["position"] is not None]
    ws.cell(row=row_sum, column=1, value="Résumé").font = Font(bold=True, size=11)
    ws.cell(row=row_sum + 1, column=1, value="Mots-clés analysés")
    ws.cell(row=row_sum + 1, column=2, value=len(results))
    ws.cell(row=row_sum + 2, column=1, value="Trouvés dans le top 100")
    ws.cell(row=row_sum + 2, column=2, value=len(found))
    ws.cell(row=row_sum + 3, column=1, value="Top 3")
    ws.cell(row=row_sum + 3, column=2, value=len([r for r in found if r["position"] <= 3]))
    ws.cell(row=row_sum + 4, column=1, value="Top 10")
    ws.cell(row=row_sum + 4, column=2, value=len([r for r in found if r["position"] <= 10]))
    if found:
        ws.cell(row=row_sum + 5, column=1, value="Position moyenne")
        avg = sum(r["position"] for r in found) / len(found)
        ws.cell(row=row_sum + 5, column=2, value=round(avg, 1))

    # Write to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"seo-positions-{domain.replace('.', '-')}-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=5001, debug=True)
